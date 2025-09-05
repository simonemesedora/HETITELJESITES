import os
import re
import unicodedata
import pdfplumber
import pandas as pd
from tkinter import Tk, filedialog
from pypdf import PdfReader, PdfWriter
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from thefuzz import process as fuzzy_process

def normalize_date_to_obj(date_str):
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%d.%m.%Y", "%m/%d/%Y", "%m/%d/%y",
                "%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except Exception:
            continue
    return None

def extract_daily_worked_dates(text):
    day_names_pattern = r"\b(Hétfő|Kedd|Szerda|Csütörtök|Péntek|Szombat|Vasárnap|Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b"
    lines = text.splitlines()
    dates = []
    for line in lines:
        if re.search(day_names_pattern, line, re.IGNORECASE):
            date_search = re.search(r"(\d{1,4}[-./]\d{1,2}[-./]\d{1,4}|\d{1,2}/\d{1,2}/\d{2,4})", line)
            if date_search:
                dt_obj = normalize_date_to_obj(date_search.group(1))
                if dt_obj:
                    dates.append(dt_obj)
    return dates

def extract_details_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            pt = page.extract_text()
            if pt:
                full_text += pt + "\n"

    name_match = re.search(r"Name:\s*(.+?)\s*Company", full_text, re.DOTALL)
    name = name_match.group(1).strip() if name_match else "Name not found"
    name = " ".join(name.split())

    daily_dates = extract_daily_worked_dates(full_text)
    if daily_dates:
        first_day = min(daily_dates).strftime("%Y/%m/%d")
        last_day = max(daily_dates).strftime("%Y/%m/%d")
    else:
        first_day, last_day = "Not found", "Not found"

    total_match = re.search(r"(Összesen|TOTAL)[:\s]*([\d\s,\.]+)Ft", full_text)
    if total_match:
        total_payment_str = total_match.group(2).replace(" ", "").replace(",", "").replace(".", "")
        total_payment = int(total_payment_str) if total_payment_str.isdigit() else 0
    else:
        total_payment = 0

    return {
        "Név": name,
        "Időszak kezdete": first_day,
        "Időszak vége": last_day,
        "Összeg": total_payment,
        "Filename": os.path.basename(pdf_path)
    }

def remove_links_from_pdfs(folder_path):
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".pdf"):
            input_path = os.path.join(folder_path, filename)
            try:
                reader = PdfReader(input_path)
            except Exception as e:
                print(f"Skipping {filename}: cannot open PDF ({e})")
                continue
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            try:
                if hasattr(writer, "remove_links"):
                    writer.remove_links()
            except KeyError:
                print(f"Warning: No annotations found in {filename}, skipping link removal for this file.")
            with open(input_path, "wb") as out_file:
                writer.write(out_file)
    print("PDF link removal completed.")

def select_folder(title):
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title=title)
    root.destroy()
    return folder_selected

def select_excel_file(title):
    root = Tk()
    root.withdraw()
    file_selected = filedialog.askopenfilename(title=title,
                                               filetypes=[("Excel files", "*.xlsx *.xls")])
    root.destroy()
    return file_selected

def normalize_name(name):
    if not isinstance(name, str):
        return ""
    name = unicodedata.normalize("NFKD", name).encode("ASCII", "ignore").decode()
    name = re.sub(r'[^a-zA-Z0-9 ]', ' ', name)
    name = name.lower().strip()
    name = re.sub(r'\s+', ' ', name)
    return name

def reverse_name(name):
    parts = name.split()
    if len(parts) > 1:
        return " ".join(parts[::-1])
    return name

def safe_min_date(series):
    dates = [normalize_date_to_obj(d) for d in series if d not in ("Not found", None, "")]
    dates = [d for d in dates if d is not None]
    return min(dates) if dates else pd.NaT

def safe_max_date(series):
    dates = [normalize_date_to_obj(d) for d in series if d not in ("Not found", None, "")]
    dates = [d for d in dates if d is not None]
    return max(dates) if dates else pd.NaT

def expand_filenames_rows(df):
    expanded_rows = []
    for _, row in df.iterrows():
        files = row['Fájlok']
        if isinstance(files, str) and "," in files and row['Név'].strip().lower() == "name not found":
            filenames = [f.strip() for f in files.split(",")]
            for f in filenames:
                new_row = row.copy()
                new_row['Fájlok'] = f
                expanded_rows.append(new_row)
        else:
            expanded_rows.append(row)
    return pd.DataFrame(expanded_rows)

def update_andrassy_with_formatting_preserved(folder, extracted_df, excel_path):
    if not os.path.exists(excel_path):
        print("ANDRASSY file not found.")
        return None

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Export"] if "Export" in wb.sheetnames else wb.active

    # Corrected header extraction
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col_idx = header.index("Név") + 1
    kezd_col_idx = header.index("Időszak kezdete") + 1
    vege_col_idx = header.index("Időszak vége") + 1
    osszeg_col_idx = header.index("Összeg") + 1

    andrassy_names = []
    andrassy_rows_map = {}
    for row in ws.iter_rows(min_row=2):
        cell_name = row[name_col_idx - 1].value
        if cell_name:
            nname = normalize_name(cell_name)
            andrassy_names.append(nname)
            andrassy_rows_map[nname] = row

    extracted_df['Név_norm'] = extracted_df['Név'].apply(normalize_name)
    grouped = extracted_df.groupby('Név_norm').agg({
        'Időszak kezdete': safe_min_date,
        'Időszak vége': safe_max_date,
        'Összeg': 'sum',
        'Név': 'first',
        'Filename': lambda f: ", ".join(sorted(set(f)))
    }).reset_index()

    matched_norm_names = set()
    names_with_multiple_periods = set()
    period_counts = extracted_df['Név_norm'].value_counts()

    for idx, ext_norm_name in enumerate(grouped['Név_norm']):
        best_match, score = fuzzy_process.extractOne(
            ext_norm_name, andrassy_names, scorer=fuzzy_process.fuzz.token_set_ratio)
        if score >= 70:
            matched_norm_names.add(ext_norm_name)
            row = andrassy_rows_map[best_match]
            extracted_row = grouped.iloc[idx]
            if kezd_col_idx and extracted_row['Időszak kezdete'] is not pd.NaT:
                row[kezd_col_idx - 1].value = (
                    extracted_row['Időszak kezdete'].strftime("%Y/%m/%d") if not pd.isna(extracted_row['Időszak kezdete']) else "")
            if vege_col_idx and extracted_row['Időszak vége'] is not pd.NaT:
                row[vege_col_idx - 1].value = (
                    extracted_row['Időszak vége'].strftime("%Y/%m/%d") if not pd.isna(extracted_row['Időszak vége']) else "")
            if osszeg_col_idx:
                row[osszeg_col_idx - 1].value = extracted_row['Összeg']
            if period_counts.get(ext_norm_name, 0) > 1:
                names_with_multiple_periods.add(ext_norm_name)

    unmatched_norm_names = set(grouped['Név_norm']) - matched_norm_names
    unmatched_rows = grouped[grouped['Név_norm'].isin(unmatched_norm_names)].copy()
    unmatched_rows.drop(columns=['Név_norm'], inplace=True)
    unmatched_rows = unmatched_rows.rename(columns={'Filename': 'Fájlok'})
    unmatched_rows = unmatched_rows[['Név', 'Időszak kezdete', 'Időszak vége', 'Összeg', 'Fájlok']]
    unmatched_rows = expand_filenames_rows(unmatched_rows)

    multi_period_rows = grouped[grouped['Név_norm'].isin(names_with_multiple_periods)].copy()
    multi_period_rows.drop(columns=['Név_norm'], inplace=True)
    multi_period_rows = multi_period_rows.rename(columns={'Filename': 'Fájlok'})
    multi_period_rows = multi_period_rows[['Név', 'Időszak kezdete', 'Időszak vége', 'Összeg', 'Fájlok']]

    if "Not Matched" in wb.sheetnames:
        wb.remove(wb["Not Matched"])
    ws_not_matched = wb.create_sheet(title="Not Matched")

    ws_not_matched.append(list(unmatched_rows.columns))
    for row_idx, row in enumerate(dataframe_to_rows(unmatched_rows, index=False, header=False), start=2):
        ws_not_matched.append(row)

    if not multi_period_rows.empty:
        ws_not_matched.append([])
        ws_not_matched.append(list(multi_period_rows.columns))
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        start_row = ws_not_matched.max_row + 1
        for i, r in enumerate(dataframe_to_rows(multi_period_rows, index=False, header=False)):
            ws_not_matched.append(r)
            row_number = start_row + i
            for col_num in range(1, len(multi_period_rows.columns) + 1):
                ws_not_matched.cell(row=row_number, column=col_num).fill = fill_green

    today_str = datetime.today().strftime('%Y%m%d')
    filename_no_ext, ext = os.path.splitext(os.path.basename(excel_path))
    output_file = f"{filename_no_ext}_{today_str}{ext}"
    output_path = os.path.join(folder, output_file)

    wb.save(output_path)
    print(f"\nUpdated ANDRASSY file saved: {output_path}")
    return output_path

if __name__ == "__main__":
    folder = select_folder("Select Folder Containing PDF Files")
    if not folder:
        print("No folder selected, exiting.")
        exit()
    if not os.path.isdir(folder):
        print("Selected path is not a valid folder, exiting.")
        exit()

    remove_links_from_pdfs(folder)

    extracted_data = []
    for filename in os.listdir(folder):
        if filename.lower().endswith(".pdf"):
            file_path = os.path.join(folder, filename)
            try:
                data = extract_details_from_pdf(file_path)
                extracted_data.append(data)
            except Exception as e:
                print(f"Skipping extraction for {filename}: {e}")

    extracted_df = pd.DataFrame(extracted_data)

    excel_path = select_excel_file("Select the Main Excel File")
    if not excel_path:
        print("No Excel file selected, exiting.")
        exit()

    update_andrassy_with_formatting_preserved(folder, extracted_df, excel_path)
